import os
import glob
import shutil
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUPPORTED_EXTENSIONS = ('.csv', '.xlsx', '.xls', '.tsv', '.txt')


def copy_matched_files(matched_file_paths, destination_folder):
    """
    Copies all matched target files to a destination folder.
    Handles filename collisions by appending index numbers.
    Returns list of copied file destination paths.
    """
    if not matched_file_paths or not destination_folder:
        return []

    os.makedirs(destination_folder, exist_ok=True)
    copied_files = []
    used_filenames = set()

    for src_path in matched_file_paths:
        if not os.path.exists(src_path):
            continue

        base_name = os.path.basename(src_path)
        name, ext = os.path.splitext(base_name)

        target_name = base_name
        counter = 1
        while target_name.lower() in used_filenames:
            target_name = f"{name}_matched_{counter}{ext}"
            counter += 1

        used_filenames.add(target_name.lower())
        dest_path = os.path.join(destination_folder, target_name)
        shutil.copy2(src_path, dest_path)
        copied_files.append(dest_path)

    return copied_files


def read_data_file(file_path, sheet_name=0):
    """
    Reads a data file (.csv, .xlsx, .xls, .tsv) into a pandas DataFrame.
    All data is loaded as string to preserve leading zeros (e.g. phone numbers, IDs).
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext == '.csv':
            # Try utf-8 first, fallback to latin-1 if encoding error
            try:
                df = pd.read_csv(file_path, dtype=str, keep_default_na=False)
            except UnicodeDecodeError:
                df = pd.read_csv(file_path, dtype=str, keep_default_na=False, encoding='latin-1')
        elif ext == '.tsv' or ext == '.txt':
            try:
                df = pd.read_csv(file_path, sep='\t', dtype=str, keep_default_na=False)
            except UnicodeDecodeError:
                df = pd.read_csv(file_path, sep='\t', dtype=str, keep_default_na=False, encoding='latin-1')
        elif ext in ('.xlsx', '.xls'):
            # Read excel sheet as strings
            df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str, keep_default_na=False)
        else:
            raise ValueError(f"Unsupported file extension: '{ext}'. Supported: {SUPPORTED_EXTENSIONS}")

        # Clean column names (strip whitespace)
        df.columns = [str(col).strip() for col in df.columns]

        # Strip strings in all cells
        for col in df.columns:
            df[col] = df[col].astype(str).str.strip()

        return df

    except Exception as e:
        raise RuntimeError(f"Error reading file '{os.path.basename(file_path)}': {str(e)}")


def find_target_files(target_input, recursive=False):
    """
    Finds all supported data files from a directory path, a single file, or a list of files.
    """
    target_files = []

    if isinstance(target_input, list):
        for item in target_input:
            target_files.extend(find_target_files(item, recursive=recursive))
        return sorted(list(set(target_files)))

    if os.path.isfile(target_input):
        ext = os.path.splitext(target_input)[1].lower()
        if ext in SUPPORTED_EXTENSIONS:
            return [os.path.abspath(target_input)]
        else:
            return []

    if os.path.isdir(target_input):
        pattern = os.path.join(target_input, "**" if recursive else "*")
        found = glob.glob(pattern, recursive=recursive)
        for fpath in found:
            if os.path.isfile(fpath):
                ext = os.path.splitext(fpath)[1].lower()
                if ext in SUPPORTED_EXTENSIONS:
                    target_files.append(os.path.abspath(fpath))

        return sorted(target_files)

    return []


def normalize_val(val, ignore_case=True):
    """
    Normalizes string values for accurate matching.
    Strips leading/trailing whitespace, single/double quotes, and normalizes case.
    """
    if val is None or pd.isna(val):
        return ""
    val_str = str(val).strip().strip("'\"").strip()
    return val_str.lower() if ignore_case else val_str


def make_composite_key(row, columns, ignore_case=True):
    """
    Generates a composite string key for matching from specified columns.
    """
    vals = [normalize_val(row.get(col, ""), ignore_case=ignore_case) for col in columns]
    return "||".join(vals)


def clean_header_name(name):
    return str(name).lower().strip().rstrip('.').strip()


def find_matching_headers(target_columns, key_columns):
    """
    Matches key columns case-insensitively against target file column headers.
    Returns a dictionary mapping source key column -> target key column name.
    """
    target_map = {clean_header_name(col): col for col in target_columns}
    mapping = {}

    for kcol in key_columns:
        k_clean = clean_header_name(kcol)
        if k_clean in target_map:
            mapping[kcol] = target_map[k_clean]

    return mapping


def compare_datasets(source_file_path, target_files, key_columns, ignore_case=True):
    """
    Compares source data against multiple target files using specified key columns.
    """
    source_name = os.path.basename(source_file_path)
    source_df = read_data_file(source_file_path)

    if source_df.empty:
        raise ValueError(f"Source file '{source_name}' is empty.")

    # Validate key columns in source
    missing_source_keys = [k for k in key_columns if k not in source_df.columns]
    if missing_source_keys:
        raise ValueError(
            f"Key column(s) {missing_source_keys} not found in source file columns: {list(source_df.columns)}"
        )

    # Initialize tracking structures for source rows
    source_records = source_df.to_dict('records')
    source_row_matches = [
        {
            'matched_files': [],
            'matched_file_count': 0,
            'match_details': []
        } for _ in range(len(source_records))
    ]

    per_file_stats = []

    # Process each target file
    for target_path in target_files:
        target_name = os.path.basename(target_path)
        try:
            target_df = read_data_file(target_path)
        except Exception as e:
            per_file_stats.append({
                'File Name': target_name,
                'File Path': target_path,
                'Status': f'ERROR: {str(e)}',
                'Total Rows': 0,
                'Matched Source Records': 0,
                'Unique Keys Matched': 0
            })
            continue

        if target_df.empty:
            per_file_stats.append({
                'File Name': target_name,
                'File Path': target_path,
                'Status': 'EMPTY FILE',
                'Total Rows': 0,
                'Matched Source Records': 0,
                'Unique Keys Matched': 0
            })
            continue

        # Map source key columns to target column headers
        header_mapping = find_matching_headers(target_df.columns, key_columns)
        if len(header_mapping) < len(key_columns):
            missing = [k for k in key_columns if k not in header_mapping]
            per_file_stats.append({
                'File Name': target_name,
                'File Path': target_path,
                'Status': f'SKIPPED (Missing key columns: {missing})',
                'Total Rows': len(target_df),
                'Matched Source Records': 0,
                'Unique Keys Matched': 0
            })
            continue

        target_key_cols = [header_mapping[k] for k in key_columns]

        # Build lookup set from target file
        target_key_set = set()
        for _, t_row in target_df.iterrows():
            t_key = make_composite_key(t_row, target_key_cols, ignore_case=ignore_case)
            target_key_set.add(t_key)

        # Match against source records
        matched_in_this_file = 0
        matched_keys_this_file = set()

        for idx, s_row in enumerate(source_records):
            s_key = make_composite_key(s_row, key_columns, ignore_case=ignore_case)
            if s_key in target_key_set:
                matched_in_this_file += 1
                matched_keys_this_file.add(s_key)
                source_row_matches[idx]['matched_files'].append(target_name)
                source_row_matches[idx]['matched_file_count'] += 1

        per_file_stats.append({
            'File Name': target_name,
            'File Path': target_path,
            'Status': 'SUCCESS',
            'Total Rows': len(target_df),
            'Matched Source Records': matched_in_this_file,
            'Unique Keys Matched': len(matched_keys_this_file)
        })

    # Collect target filepaths that matched at least 1 source record
    per_file_df = pd.DataFrame(per_file_stats)
    matched_target_paths = []
    if not per_file_df.empty and 'Matched Source Records' in per_file_df.columns:
        matched_rows = per_file_df[per_file_df['Matched Source Records'] > 0]
        matched_target_paths = matched_rows['File Path'].tolist()

    # Build final result dataframes
    source_results_list = []
    total_matched_sources = 0

    for idx, s_row in enumerate(source_records):
        match_info = source_row_matches[idx]
        is_matched = match_info['matched_file_count'] > 0
        if is_matched:
            total_matched_sources += 1

        res_row = dict(s_row)
        res_row['MATCH_STATUS'] = 'FOUND' if is_matched else 'NOT FOUND'
        res_row['MATCH_COUNT'] = match_info['matched_file_count']
        res_row['MATCHED_FILES'] = ", ".join(match_info['matched_files']) if match_info['matched_files'] else "None"
        source_results_list.append(res_row)

    full_results_df = pd.DataFrame(source_results_list)
    matched_df = full_results_df[full_results_df['MATCH_STATUS'] == 'FOUND'].copy()
    unmatched_df = full_results_df[full_results_df['MATCH_STATUS'] == 'NOT FOUND'].copy()

    total_source_count = len(source_df)
    total_unmatched_sources = total_source_count - total_matched_sources
    match_percentage = (total_matched_sources / total_source_count * 100) if total_source_count > 0 else 0.0

    summary_stats = {
        'Source File Name': source_name,
        'Source File Path': os.path.abspath(source_file_path),
        'Total Target Files Checked': len(target_files),
        'Matched Target Files Count': len(matched_target_paths),
        'Key Columns Used': ", ".join(key_columns),
        'Total Source Records': total_source_count,
        'Found in Targets (Matched)': total_matched_sources,
        'Missing in Targets (Unmatched)': total_unmatched_sources,
        'Match Percentage': f"{match_percentage:.2f}%"
    }

    return {
        'summary_stats': summary_stats,
        'per_file_stats': per_file_df,
        'full_results_df': full_results_df,
        'matched_df': matched_df,
        'unmatched_df': unmatched_df,
        'matched_target_paths': matched_target_paths
    }


def export_to_excel(results, output_file_path):
    """
    Exports comparison results to an Excel workbook with formatting and multiple tabs.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styling definitions
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    found_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Light Green
    found_font = Font(name="Calibri", size=11, color="375623", bold=True)
    
    not_found_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Light Red
    not_found_font = Font(name="Calibri", size=11, color="C65911", bold=True)

    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    bold_font = Font(name="Calibri", size=11, bold=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # -------------------------------------------------------------
    # Sheet 1: Dashboard Summary
    # -------------------------------------------------------------
    ws_dash = wb.create_sheet(title="Dashboard Summary")
    ws_dash.views.sheetView[0].showGridLines = True

    ws_dash.cell(row=1, column=1, value="DATA COMPARISON REPORT DASHBOARD").font = title_font
    
    ws_dash.cell(row=3, column=1, value="Metric").font = header_font
    ws_dash.cell(row=3, column=1).fill = header_fill
    ws_dash.cell(row=3, column=2, value="Value").font = header_font
    ws_dash.cell(row=3, column=2).fill = header_fill

    row_idx = 4
    for key, val in results['summary_stats'].items():
        cell_k = ws_dash.cell(row=row_idx, column=1, value=key)
        cell_k.font = bold_font
        cell_k.border = thin_border
        
        cell_v = ws_dash.cell(row=row_idx, column=2, value=str(val))
        cell_v.border = thin_border
        row_idx += 1

    # -------------------------------------------------------------
    # Sheet 2: Target File Breakdown
    # -------------------------------------------------------------
    ws_targets = wb.create_sheet(title="Target Files Breakdown")
    ws_targets.views.sheetView[0].showGridLines = True
    
    per_file_df = results['per_file_stats']
    if not per_file_df.empty:
        # Write headers
        for c_idx, col_name in enumerate(per_file_df.columns, start=1):
            cell = ws_targets.cell(row=1, column=c_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for r_idx, row in per_file_df.iterrows():
            for c_idx, col_name in enumerate(per_file_df.columns, start=1):
                val = row[col_name]
                cell = ws_targets.cell(row=r_idx + 2, column=c_idx, value=val)
                cell.border = thin_border
                if col_name == 'Status':
                    if val == 'SUCCESS':
                        cell.fill = found_fill
                        cell.font = found_font
                    elif 'SKIPPED' in str(val) or 'ERROR' in str(val):
                        cell.fill = not_found_fill
                        cell.font = not_found_font

    # -------------------------------------------------------------
    # Helper to write dataframes to Excel sheet
    # -------------------------------------------------------------
    def write_df_to_sheet(sheet_title, df):
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        
        if df.empty:
            ws.cell(row=1, column=1, value="No records found.").font = bold_font
            return ws

        # Write Headers
        for c_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write Data
        for r_idx, row in df.iterrows():
            curr_row = r_idx + 2
            for c_idx, col_name in enumerate(df.columns, start=1):
                val = row[col_name]
                cell = ws.cell(row=curr_row, column=c_idx, value=str(val))
                cell.border = thin_border

                # Format MATCH_STATUS
                if col_name == 'MATCH_STATUS':
                    cell.alignment = Alignment(horizontal="center")
                    if val == 'FOUND':
                        cell.fill = found_fill
                        cell.font = found_font
                    else:
                        cell.fill = not_found_fill
                        cell.font = not_found_font

        return ws

    # -------------------------------------------------------------
    # Sheet 3: Full Comparison Results
    # -------------------------------------------------------------
    write_df_to_sheet("All Source Records", results['full_results_df'])

    # Sheet 4: Matched Source Records
    write_df_to_sheet("Matched Source Records", results['matched_df'])

    # Sheet 5: Unmatched Source Records
    write_df_to_sheet("Unmatched Source Records", results['unmatched_df'])

    # Auto-adjust column widths across all worksheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.row == 1 and sheet.title == "Dashboard Summary":
                    continue # Skip title line
                max_len = max(max_len, len(val_str))
            sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

    # Save workbook
    os.makedirs(os.path.dirname(os.path.abspath(output_file_path)), exist_ok=True)
    wb.save(output_file_path)
    wb.close()
    return output_file_path
